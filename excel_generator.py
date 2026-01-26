"""
Generador de archivos Excel con catálogos de productos.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime
import os


class ExcelGenerator:
    """Genera archivos Excel con catálogos de productos."""
    
    def __init__(self, output_dir: str = "./output"):
        """
        Inicializa el generador de Excel.
        
        Args:
            output_dir: Directorio de salida para los archivos
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_catalog(self, products: List[Dict[str, str]], domain: str) -> str:
        """
        Genera un archivo Excel con el catálogo de productos.
        
        Args:
            products: Lista de productos
            domain: Dominio del sitio web
            
        Returns:
            Ruta del archivo generado
        """
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo"
        
        # Configurar encabezados
        headers = ['nombre_articulo', 'precio', 'descripcion', 'url_imagenes']
        self._write_headers(ws, headers)
        
        # Escribir datos de productos
        for idx, product in enumerate(products, start=2):
            ws.cell(row=idx, column=1, value=product.get('nombre_articulo', ''))
            ws.cell(row=idx, column=2, value=product.get('precio', ''))
            ws.cell(row=idx, column=3, value=product.get('descripcion', ''))
            ws.cell(row=idx, column=4, value=product.get('url_imagenes', ''))
        
        # Ajustar anchos de columna
        self._auto_adjust_columns(ws)
        
        # Agregar filtros
        ws.auto_filter.ref = ws.dimensions
        
        # Agregar hoja de resumen
        self._add_summary_sheet(wb, products, domain)
        
        # Guardar archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"catalogo_{self._sanitize_domain(domain)}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb.save(filepath)
        return filepath
    
    def _write_headers(self, ws, headers: List[str]):
        """
        Escribe los encabezados con formato.
        
        Args:
            ws: Worksheet de openpyxl
            headers: Lista de nombres de columnas
        """
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
    
    def _auto_adjust_columns(self, ws):
        """
        Ajusta automáticamente el ancho de las columnas.
        
        Args:
            ws: Worksheet de openpyxl
        """
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            # Limitar el ancho máximo
            adjusted_width = min(length + 2, 60)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
    
    def _add_summary_sheet(self, wb, products: List[Dict], domain: str):
        """
        Agrega una hoja de resumen con estadísticas.
        
        Args:
            wb: Workbook de openpyxl
            products: Lista de productos
            domain: Dominio del sitio
        """
        ws = wb.create_sheet("Resumen")
        
        # Título
        ws.cell(row=1, column=1, value="Resumen del Catálogo").font = Font(bold=True, size=14)
        
        # Información general
        ws.cell(row=3, column=1, value="Sitio web:")
        ws.cell(row=3, column=2, value=domain)
        
        ws.cell(row=4, column=1, value="Fecha de extracción:")
        ws.cell(row=4, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        ws.cell(row=5, column=1, value="Total de productos:")
        ws.cell(row=5, column=2, value=len(products))
        
        # Estadísticas de precios
        prices = []
        for product in products:
            price_str = product.get('precio', '')
            # Intentar extraer números del precio
            import re
            numbers = re.findall(r'[\d,\.]+', price_str)
            if numbers:
                try:
                    # Convertir a float (manejar comas y puntos)
                    price_num = float(numbers[0].replace(',', ''))
                    prices.append(price_num)
                except ValueError:
                    pass
        
        if prices:
            ws.cell(row=7, column=1, value="Precio mínimo:")
            ws.cell(row=7, column=2, value=f"${min(prices):,.2f}")
            
            ws.cell(row=8, column=1, value="Precio máximo:")
            ws.cell(row=8, column=2, value=f"${max(prices):,.2f}")
            
            ws.cell(row=9, column=1, value="Precio promedio:")
            ws.cell(row=9, column=2, value=f"${sum(prices)/len(prices):,.2f}")
        
        # Productos con imágenes
        products_with_images = sum(1 for p in products if p.get('url_imagenes'))
        ws.cell(row=11, column=1, value="Productos con imágenes:")
        ws.cell(row=11, column=2, value=products_with_images)
        
        # Productos con descripción
        products_with_desc = sum(1 for p in products if p.get('descripcion'))
        ws.cell(row=12, column=1, value="Productos con descripción:")
        ws.cell(row=12, column=2, value=products_with_desc)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    @staticmethod
    def _sanitize_domain(domain: str) -> str:
        """
        Limpia el nombre de dominio para usar en nombres de archivo.
        
        Args:
            domain: Nombre de dominio
            
        Returns:
            Dominio sanitizado
        """
        # Eliminar protocolo
        domain = domain.replace('https://', '').replace('http://', '')
        # Eliminar www
        domain = domain.replace('www.', '')
        # Eliminar barra final
        domain = domain.rstrip('/')
        # Reemplazar caracteres especiales
        domain = domain.replace('/', '_').replace(':', '_')
        return domain
